"""Lit les fichiers AOH, MARSA MAROC, ANP, INPPLC pour voir leur structure complète"""
import openpyxl, xlrd, os

fichiers = {
    "AOH": r"c:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\AOH\MD HAO.xlsx",
    "MARSA MAROC": r"c:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\MARSA MAROC\MP MM .xlsx",
    "INPPLC": r"c:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\INPPLC\Masters Data INPPLC.XLS",
}

for nom, filepath in fichiers.items():
    print(f"\n{'='*50}")
    print(f"MARCHÉ : {nom}")
    print(f"{'='*50}")
    ext = filepath.lower().split(".")[-1]
    try:
        if ext == "xlsx":
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                print(f"\n  Feuille : {sheet_name}")
                ws = wb[sheet_name]
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    row_clean = [str(c).strip() if c is not None else "" for c in row]
                    if any(row_clean):
                        print(f"    L{i+1}: {row_clean[:8]}")
                    if i > 30:
                        print("    ...")
                        break
            wb.close()
        else:  # xls
            wb = xlrd.open_workbook(filepath)
            for sheet_name in wb.sheet_names():
                print(f"\n  Feuille : {sheet_name}")
                ws = wb.sheet_by_name(sheet_name)
                for i in range(min(35, ws.nrows)):
                    row = [str(ws.cell_value(i, j)).strip() for j in range(min(8, ws.ncols))]
                    if any(row):
                        print(f"    L{i+1}: {row}")
    except Exception as e:
        print(f"  ERREUR: {e}")
