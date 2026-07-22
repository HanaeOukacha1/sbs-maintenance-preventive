"""
Script d'analyse des Master Data SBS
Lit tous les fichiers Excel/XLS du dossier MASTER DATA
et génère un rapport sur la structure des données.
"""
import os
import json
import sys

# On installe openpyxl et xlrd si besoin
try:
    import openpyxl
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl xlrd -q")
    import openpyxl

try:
    import xlrd
except ImportError:
    os.system(f"{sys.executable} -m pip install xlrd -q")
    import xlrd

MASTER_DATA_PATH = r"c:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA"
OUTPUT_PATH = r"c:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\analyse_rapport.json"

rapport = {}

print("=" * 60)
print("ANALYSE DES MASTER DATA SBS")
print("=" * 60)

for marche_nom in sorted(os.listdir(MASTER_DATA_PATH)):
    marche_path = os.path.join(MASTER_DATA_PATH, marche_nom)
    if not os.path.isdir(marche_path):
        continue

    print(f"\n📁 MARCHÉ : {marche_nom}")
    rapport[marche_nom] = {"fichiers": []}

    for fichier in sorted(os.listdir(marche_path)):
        fichier_path = os.path.join(marche_path, fichier)
        ext = fichier.lower().split(".")[-1]

        info_fichier = {"nom": fichier, "type": ext, "feuilles": []}

        if ext in ("xlsx", "xls"):
            try:
                if ext == "xlsx":
                    wb = openpyxl.load_workbook(fichier_path, read_only=True, data_only=True)
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        # Lire les 3 premières lignes pour comprendre la structure
                        rows = []
                        for i, row in enumerate(ws.iter_rows(values_only=True)):
                            if i >= 5:
                                break
                            row_clean = [str(c).strip() if c is not None else "" for c in row]
                            if any(row_clean):
                                rows.append(row_clean)
                        info_fichier["feuilles"].append({
                            "nom": sheet_name,
                            "apercu_lignes": rows
                        })
                    wb.close()

                else:  # .xls
                    wb = xlrd.open_workbook(fichier_path)
                    for sheet_name in wb.sheet_names():
                        ws = wb.sheet_by_name(sheet_name)
                        rows = []
                        for i in range(min(5, ws.nrows)):
                            row = [str(ws.cell_value(i, j)).strip() for j in range(ws.ncols)]
                            if any(row):
                                rows.append(row)
                        info_fichier["feuilles"].append({
                            "nom": sheet_name,
                            "apercu_lignes": rows
                        })

                print(f"  ✅ {fichier} → {len(info_fichier['feuilles'])} feuille(s)")
                for f in info_fichier["feuilles"]:
                    print(f"     📄 Feuille: '{f['nom']}'")
                    if f["apercu_lignes"]:
                        print(f"        1ère ligne: {f['apercu_lignes'][0][:6]}")

            except Exception as e:
                info_fichier["erreur"] = str(e)
                print(f"  ⚠️  {fichier} → ERREUR: {e}")

        elif ext in ("docx",):
            print(f"  📝 {fichier} → Document Word (checklist papier)")
            info_fichier["note"] = "Checklist Word — à analyser manuellement"

        rapport[marche_nom]["fichiers"].append(info_fichier)

# Sauvegarde du rapport JSON
with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
    json.dump(rapport, f, ensure_ascii=False, indent=2)

print("\n" + "=" * 60)
print(f"✅ Rapport sauvegardé : {OUTPUT_PATH}")
print("=" * 60)
