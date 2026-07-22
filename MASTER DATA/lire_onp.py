"""Lit le fichier ONP complet pour extraire les sites et colonnes"""
import openpyxl
import sys, os

filepath = r"c:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\ONP\MP ONP.xlsx"
wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

for sheet_name in wb.sheetnames:
    print(f"\n=== FEUILLE : {sheet_name} ===")
    ws = wb[sheet_name]
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        row_clean = [str(c).strip() if c is not None else "" for c in row]
        if any(row_clean):
            print(f"  Ligne {i+1}: {row_clean[:10]}")  # Max 10 colonnes
        if i > 100:  # Limite pour ne pas tout afficher
            print("  ... (tronqué à 100 lignes)")
            break
wb.close()
