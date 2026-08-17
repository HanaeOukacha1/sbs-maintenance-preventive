# coding: utf-8
import sys
import openpyxl
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(
    r'C:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\AMEE\MP AMEE MARRAKECH 1T-2026.xlsx',
    data_only=True
)

# Feuilles: ['DATA CENTER', 'Feuil9', 'UC ', 'MISE A JOUR', 'IMPRIMANTE ET MFP ', 'Feuil1']

for sheet_name in ['DATA CENTER', 'UC ', 'MISE A JOUR', 'IMPRIMANTE ET MFP ']:
    ws = wb[sheet_name]
    print(f"\n{'='*50}")
    print(f"Feuille: '{sheet_name}' | {ws.max_row} lignes | {ws.max_column} colonnes")
    # Show ALL rows up to 40
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), 1):
        vals = [str(v)[:25] if v is not None else '' for v in row[:12]]
        if any(v.strip() for v in vals):
            print(f"  L{i:2d}: {' | '.join(vals)}")
