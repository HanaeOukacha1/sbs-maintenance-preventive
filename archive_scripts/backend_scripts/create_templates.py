import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
import os

master_dir = r'C:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\MSANTE'
template_dir = r'C:\Users\hanae\Desktop\Stage PFA 2026\backend\app\templates'

files_to_convert = ['CAPM S2.XLS', 'DPRF S2.XLS']

bold_font = Font(bold=True)
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

for file in files_to_convert:
    in_path = os.path.join(master_dir, file)
    out_name = f"template_msante_{file.split(' ')[0].lower()}.xlsx"
    out_path = os.path.join(template_dir, out_name)
    
    print(f"Creating {out_path} from {in_path}...")
    xls = pd.ExcelFile(in_path)
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # remove default sheet
    
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
        
        ws = wb.create_sheet(title=str(sheet_name)[:31])
        
        # We only want rows 0 to 8 (which is rows 1 to 9 in excel)
        # We will write them as headers
        header_df = df.head(9)
        
        for r_idx, row in header_df.iterrows():
            for c_idx, val in enumerate(row):
                if pd.notna(val):
                    cell = ws.cell(row=r_idx+1, column=c_idx+1, value=str(val))
                    cell.font = bold_font
                    
        # Apply borders to row 9 (the column headers)
        for c in range(1, 8):
            cell = ws.cell(row=9, column=c)
            cell.border = thin_border
            cell.font = bold_font
            
        # Set some column widths
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        
    wb.save(out_path)
print("Conversion done without COM!")
