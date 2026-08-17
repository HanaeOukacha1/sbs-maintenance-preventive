# coding: utf-8
import sys
import openpyxl
sys.stdout.reconfigure(encoding='utf-8')

def inspect_file(path, label):
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"\n{'='*60}")
    print(f"{label}")
    print(f"Feuilles: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row

        print(f"\n--- Feuille: '{sheet_name}' | {max_row} lignes ---")
        
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
            vals = [str(v)[:22] if v is not None else '' for v in row[:12]]
            if any(v.strip() for v in vals):
                print(f"  L{i:2d}: {' | '.join(vals)}")
        
        if max_row > 20:
            print(f"  ... ({max_row - 20} lignes de plus)")

inspect_file(
    r'C:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\AMEE\MP AMEE MARRAKECH 1T-2026.xlsx',
    'AMEE MARRAKECH'
)
inspect_file(
    r'C:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\AMEE\NV MP AMEE RABAT 1T-2026.xlsx',
    'AMEE RABAT'
)
