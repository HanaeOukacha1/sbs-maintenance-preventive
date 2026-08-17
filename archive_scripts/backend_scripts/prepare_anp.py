import openpyxl
import os

src = r'C:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\ANP\MP ANP.xlsx'
dest = r'C:\Users\hanae\Desktop\Stage PFA 2026\backend\app\templates\template_anp.xlsx'

wb = openpyxl.load_workbook(src)

# We want to keep only the first sheet, clear data from row 7 downwards
if 'Feuil2' in wb.sheetnames:
    del wb['Feuil2']

# Rename the first sheet to something generic, or leave it.
sheet = wb.worksheets[0]
sheet.title = "Export"

# Delete rows from 7 to max
sheet.delete_rows(7, sheet.max_row)

wb.save(dest)
print(f"Created {dest}")
