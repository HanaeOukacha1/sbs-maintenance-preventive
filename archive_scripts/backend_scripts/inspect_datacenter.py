# coding: utf-8
import sys
import openpyxl
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(
    r'C:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\AMEE\MP AMEE MARRAKECH 1T-2026.xlsx',
    data_only=True
)

ws = wb['DATA CENTER']
print(f"=== MARRAKECH DATA CENTER ===")
count = 0
for row in ws.iter_rows(min_row=8, values_only=True):
    v = [str(c) if c else '' for c in row[:5]]
    if any(v):
        print(" | ".join(v))
        count += 1
print(f"Total rows: {count}")

wb2 = openpyxl.load_workbook(
    r'C:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\AMEE\NV MP AMEE RABAT 1T-2026.xlsx',
    data_only=True
)
print(f"\n=== RABAT DATA CENTER ===")
if 'DATA CENTER' in wb2.sheetnames:
    ws2 = wb2['DATA CENTER']
    count = 0
    for row in ws2.iter_rows(min_row=8, values_only=True):
        v = [str(c) if c else '' for c in row[:5]]
        if any(v):
            print(" | ".join(v))
            count += 1
    print(f"Total rows: {count}")
