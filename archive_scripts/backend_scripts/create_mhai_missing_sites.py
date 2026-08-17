# coding: utf-8
from sqlalchemy import create_engine, text
import openpyxl

engine = create_engine('mysql+pymysql://root:@localhost/sbs_db', pool_pre_ping=True)

wb = openpyxl.load_workbook(
    r'C:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\MHAI\MP HABOUS S2-24.xlsx',
    data_only=True
)

# Map sheet name -> city
SHEETS = {
    'Tanger': 'Tanger',
    'Marrakech,': 'Marrakech',
    'CASA': 'Casablanca',
}

with engine.connect() as conn:
    # Get the MHAI marche id
    marche = conn.execute(text("SELECT id FROM marches WHERE nom LIKE '%MHAI%' OR nom LIKE '%Habous%' LIMIT 1")).fetchone()
    if not marche:
        marche = conn.execute(text("SELECT id FROM marches WHERE nom LIKE '%MHAI%' LIMIT 1")).fetchone()
    # fallback: get marche from existing MHAI site
    marche_row = conn.execute(text(
        "SELECT s.marche_id FROM sites s WHERE s.checklist_type = 'MHAI' LIMIT 1"
    )).fetchone()
    marche_id = marche_row[0] if marche_row else None
    print(f'MHAI marche_id: {marche_id}')

    # Get technicien hanae id
    tech = conn.execute(text("SELECT id FROM users WHERE email = 'hanae@sbs.ma'")).fetchone()
    tech_id = tech[0] if tech else 4
    print(f'Tech id: {tech_id}')

    for sheet_name, city in SHEETS.items():
        ws = wb[sheet_name]
        
        # Find header row (row with N°, Matériel, etc.)
        header_row = None
        for i, row in enumerate(ws.iter_rows(max_row=15, values_only=True), 1):
            vals = [str(v).lower() if v else '' for v in row]
            if any('mat' in v or 'marque' in v for v in vals):
                header_row = i
                headers = row
                break
        
        if not header_row:
            print(f'  {sheet_name}: header not found')
            continue
        
        print(f'\n=== {sheet_name} (city: {city}, header at row {header_row}) ===')
        print(f'  Headers: {headers}')
        
        # Map header -> column index
        col_map = {}
        for idx, h in enumerate(headers):
            if h is None: continue
            hl = str(h).lower()
            if 'mat' in hl: col_map['materiel'] = idx
            elif 'marque' in hl: col_map['marque'] = idx
            elif 'mod' in hl: col_map['modele'] = idx
            elif 'serie' in hl or 'série' in hl: col_map['numero_serie'] = idx
            elif 'invent' in hl: col_map['numero_inventaire'] = idx
        print(f'  Col map: {col_map}')
        
        # Create site
        site_nom = f'MHAI {city}'
        existing = conn.execute(text(f"SELECT id FROM sites WHERE nom = :n"), {'n': site_nom}).fetchone()
        if existing:
            site_id = existing[0]
            print(f'  Site already exists: {site_id}')
        else:
            conn.execute(text(
                "INSERT INTO sites (nom, ville, marche_id, checklist_type, feuilles) "
                "VALUES (:nom, :ville, :marche_id, 'MHAI', NULL)"
            ), {'nom': site_nom, 'ville': city, 'marche_id': marche_id})
            conn.commit()
            site_id = conn.execute(text("SELECT LAST_INSERT_ID()")).fetchone()[0]
            print(f'  Created site {site_id}: {site_nom}')
        
        # Create mission
        existing_m = conn.execute(text(f"SELECT id FROM missions WHERE site_id = :s"), {'s': site_id}).fetchone()
        if existing_m:
            mission_id = existing_m[0]
            conn.execute(text("UPDATE missions SET technicien_id = :t, statut = 'PLANIFIEE' WHERE id = :mid"), {'t': tech_id, 'mid': mission_id})
            conn.commit()
            print(f'  Mission already exists: {mission_id} (updated)')
        else:
            conn.execute(text(
                "INSERT INTO missions (titre, site_id, technicien_id, statut, date_planifiee) "
                "VALUES (:titre, :site_id, :tech_id, 'PLANIFIEE', CURDATE())"
            ), {'titre': f'MP MHAI {city}', 'site_id': site_id, 'tech_id': tech_id})
            conn.commit()
            mission_id = conn.execute(text("SELECT LAST_INSERT_ID()")).fetchone()[0]
            print(f'  Created mission {mission_id}')
        
        # Import equipment
        existing_eq = conn.execute(text(f"SELECT COUNT(*) FROM equipements WHERE site_id = {site_id}")).fetchone()
        if existing_eq[0] > 0:
            print(f'  Equipment already imported: {existing_eq[0]} items. Skipping.')
            continue
        
        count_inserted = 0
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if all(v is None for v in row):
                continue
            materiel = row[col_map['materiel']] if 'materiel' in col_map else None
            marque = row[col_map['marque']] if 'marque' in col_map else None
            modele = row[col_map['modele']] if 'modele' in col_map else None
            numero_serie = row[col_map['numero_serie']] if 'numero_serie' in col_map else None
            numero_inventaire = row[col_map['numero_inventaire']] if 'numero_inventaire' in col_map else None
            
            if numero_serie is None and marque is None and materiel is None:
                continue
            
            # Determine type from materiel
            mat_lower = str(materiel or '').lower()
            if 'imprimante' in mat_lower or 'mfp' in mat_lower or 'photocopieur' in mat_lower:
                type_eq = 'IMPRIMANTE'
            elif 'portable' in mat_lower or 'laptop' in mat_lower:
                type_eq = 'PORTABLE'
            elif 'pc' in mat_lower or 'fixe' in mat_lower:
                type_eq = 'PC'
            else:
                type_eq = 'AUTRE'
            
            conn.execute(text(
                "INSERT INTO equipements (site_id, famille, marque, modele, numero_serie, numero_inventaire, type_equipement, is_active) "
                "VALUES (:site_id, :famille, :marque, :modele, :numero_serie, :numero_inventaire, :type_eq, 1)"
            ), {
                'site_id': site_id,
                'famille': str(materiel)[:100] if materiel else None,
                'marque': str(marque)[:100] if marque else None,
                'modele': str(modele)[:100] if modele else None,
                'numero_serie': str(numero_serie)[:100] if numero_serie else None,
                'numero_inventaire': str(numero_inventaire)[:100] if numero_inventaire else None,
                'type_eq': type_eq
            })
            count_inserted += 1
        
        conn.commit()
        print(f'  Inserted {count_inserted} equipements for {city}')

print('\nAll done!')
