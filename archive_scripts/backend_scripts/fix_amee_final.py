# coding: utf-8
"""
Script de reimport AMEE Marrakech et Rabat avec la structure exacte de chaque feuille.
"""
from sqlalchemy import create_engine, text
import openpyxl
import json

engine = create_engine('mysql+pymysql://root:@localhost/sbs_db', pool_pre_ping=True)

def val(row, idx, max_len=150):
    if idx < len(row) and row[idx] is not None:
        v = str(row[idx]).strip()
        if v.lower() in ('', 'none', 'nan'):
            return None
        return v[:max_len]
    return None

# ────────────────────────────────────────────────────────────────────────────
# MARRAKECH
# ────────────────────────────────────────────────────────────────────────────
def import_marrakech(conn):
    site = conn.execute(text("SELECT id FROM sites WHERE checklist_type = 'AMEE_MARRAKECH'")).fetchone()
    if not site:
        print("Site AMEE Marrakech introuvable")
        return
    site_id = site[0]
    
    # Clear equipements
    conn.execute(text(f"DELETE FROM equipements WHERE site_id = {site_id}"))
    conn.commit()
    
    wb = openpyxl.load_workbook(
        r'C:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\AMEE\MP AMEE MARRAKECH 1T-2026.xlsx',
        data_only=True
    )
    
    total = 0
    
    # ── 1. DATA CENTER ── (header row 7, cols: Type, Marque, Modele, N°Série)
    ws = wb['DATA CENTER']
    count = 0
    for row in ws.iter_rows(min_row=8, values_only=True):
        type_eq = val(row, 0)
        marque   = val(row, 1)
        modele   = val(row, 2)
        serie    = val(row, 3)
        if not type_eq or type_eq.lower().startswith('pour sub'):
            continue
        nom = type_eq
        teq = 'RESEAU' if any(k in str(type_eq).upper() for k in ('SERVEUR', 'SWITCH', 'KVM', 'BAIE')) else 'AUTRE'
        conn.execute(text(
            "INSERT INTO equipements (site_id, sous_site, nom, famille, marque, modele, numero_serie, type_equipement, is_active) "
            "VALUES (:sid, 'DATA CENTER', :nom, :fam, :marq, :mod, :ser, :teq, 1)"
        ), {'sid': site_id, 'nom': nom, 'fam': type_eq, 'marq': marque, 'mod': modele, 'ser': serie, 'teq': teq})
        count += 1
    conn.commit()
    print(f"  DATA CENTER: {count} équipements")
    total += count
    
    # ── 2. UC ── (header row 6, cols: Utilisateur, Type, Marque, Modele, S/N, CPU, RAM, OS, N°Inv)
    ws = wb['UC ']
    count = 0
    for row in ws.iter_rows(min_row=7, values_only=True):
        utilisateur = val(row, 0)
        type_eq     = val(row, 1)
        marque      = val(row, 2)
        modele      = val(row, 3)
        serie       = val(row, 4)
        cpu         = val(row, 5)
        ram         = val(row, 6)
        os_val      = val(row, 7)
        inventaire  = val(row, 8)
        if not type_eq:
            continue
        nom = f"{type_eq} - {utilisateur}" if utilisateur else type_eq
        teq = 'PC'
        if 'ECRAN' in str(type_eq).upper():
            teq = 'ECRAN'
        elif 'IMPRIMANTE' in str(type_eq).upper():
            teq = 'IMPRIMANTE'
        conn.execute(text(
            "INSERT INTO equipements (site_id, sous_site, nom, famille, utilisateur_nom, marque, modele, numero_serie, "
            "cpu, ram, systeme_exploitation, numero_inventaire, type_equipement, is_active) "
            "VALUES (:sid, 'UC', :nom, :fam, :user, :marq, :mod, :ser, :cpu, :ram, :os, :inv, :teq, 1)"
        ), {'sid': site_id, 'nom': nom, 'fam': type_eq, 'user': utilisateur,
            'marq': marque, 'mod': modele, 'ser': serie, 'cpu': cpu, 'ram': ram,
            'os': os_val, 'inv': inventaire, 'teq': teq})
        count += 1
    conn.commit()
    print(f"  UC: {count} équipements")
    total += count
    
    # ── 3. MISE A JOUR ── (header row 6, cols: Personnes, TYPES, Nettoyage, Fichiers temp, MAJ Windows)
    # C'est une liste de personnes + résultats de maintenance (pas d'équipements physiques)
    # On les importe comme entrées avec utilisateur_nom et famille = TYPES
    ws = wb['MISE A JOUR']
    count = 0
    for row in ws.iter_rows(min_row=7, values_only=True):
        personne = val(row, 0)
        types    = val(row, 1)
        nettoyage = val(row, 2)
        fichiers  = val(row, 3)
        maj_win   = val(row, 4)
        if not personne or personne.lower().startswith('pour sub'):
            continue
        nom = f"{types or 'PC'} - {personne}"
        # Store as JSON in description for display
        desc = f"Nettoyage: {nettoyage or '-'} | Fichiers temp: {fichiers or '-'} | MAJ Windows: {maj_win or '-'}"
        conn.execute(text(
            "INSERT INTO equipements (site_id, sous_site, nom, famille, utilisateur_nom, description, type_equipement, is_active) "
            "VALUES (:sid, 'MISE A JOUR', :nom, :fam, :user, :desc, 'PC', 1)"
        ), {'sid': site_id, 'nom': nom, 'fam': types, 'user': personne, 'desc': desc})
        count += 1
    conn.commit()
    print(f"  MISE A JOUR: {count} entrées")
    total += count
    
    # ── 4. IMPRIMANTE ET MFP ── (header row 7, cols: Type, Modèle, Emplacement, N°Série, Statut)
    ws = wb['IMPRIMANTE ET MFP ']
    count = 0
    for row in ws.iter_rows(min_row=8, values_only=True):
        type_eq    = val(row, 0)
        modele     = val(row, 1)
        emplacement = val(row, 2)
        serie      = val(row, 3)
        statut     = val(row, 4)
        if not type_eq or type_eq.lower().startswith('pour sub'):
            continue
        nom = f"{type_eq} - {modele}" if modele else type_eq
        conn.execute(text(
            "INSERT INTO equipements (site_id, sous_site, nom, famille, modele, emplacement, numero_serie, description, type_equipement, is_active) "
            "VALUES (:sid, 'IMPRIMANTE ET MFP', :nom, :fam, :mod, :emp, :ser, :desc, 'IMPRIMANTE', 1)"
        ), {'sid': site_id, 'nom': nom, 'fam': type_eq, 'mod': modele,
            'emp': emplacement, 'ser': serie, 'desc': statut})
        count += 1
    conn.commit()
    print(f"  IMPRIMANTE ET MFP: {count} équipements")
    total += count
    
    feuilles = ['DATA CENTER', 'UC', 'MISE A JOUR', 'IMPRIMANTE ET MFP']
    conn.execute(text("UPDATE sites SET feuilles = :f WHERE id = :sid"),
                 {'f': json.dumps(feuilles), 'sid': site_id})
    conn.commit()
    print(f"  => Total Marrakech: {total} entrées | Feuilles: {feuilles}")


# ────────────────────────────────────────────────────────────────────────────
# RABAT
# ────────────────────────────────────────────────────────────────────────────
def import_rabat(conn):
    site = conn.execute(text("SELECT id FROM sites WHERE checklist_type = 'AMEE_RABAT'")).fetchone()
    if not site:
        print("Site AMEE Rabat introuvable")
        return
    site_id = site[0]
    
    # Clear equipements
    conn.execute(text(f"DELETE FROM equipements WHERE site_id = {site_id}"))
    conn.commit()
    
    wb = openpyxl.load_workbook(
        r'C:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\AMEE\NV MP AMEE RABAT 1T-2026.xlsx',
        data_only=True
    )
    
    total = 0
    
    # ── 1. PC (feuille 'PC.') ── 
    # Header row 6, cols: Personnes, Type, Marque, Désigniation, Système exp, espace DD, RAM, espace libre, CPU, N°série, MAJ, ManageEngine
    ws = wb['PC.']
    count = 0
    for row in ws.iter_rows(min_row=7, values_only=True):
        utilisateur = val(row, 0)
        type_eq     = val(row, 1)
        marque      = val(row, 2)
        modele      = val(row, 3)
        os_val      = val(row, 4)
        disque      = val(row, 5)
        ram         = val(row, 6)
        cpu         = val(row, 8)
        serie       = val(row, 9)
        if not type_eq:
            continue
        if type_eq.lower().startswith('pour sub'):
            continue
        nom = f"{type_eq} - {utilisateur}" if utilisateur else type_eq
        teq = 'PC'
        if 'ECRAN' in str(type_eq).upper():
            teq = 'ECRAN'
        conn.execute(text(
            "INSERT INTO equipements (site_id, sous_site, nom, famille, utilisateur_nom, marque, modele, numero_serie, "
            "cpu, ram, disque_dur, systeme_exploitation, type_equipement, is_active) "
            "VALUES (:sid, 'PC', :nom, :fam, :user, :marq, :mod, :ser, :cpu, :ram, :dd, :os, :teq, 1)"
        ), {'sid': site_id, 'nom': nom, 'fam': type_eq, 'user': utilisateur,
            'marq': marque, 'mod': modele, 'ser': serie, 'cpu': cpu, 'ram': ram,
            'dd': disque, 'os': os_val, 'teq': teq})
        count += 1
    conn.commit()
    print(f"  PC: {count} équipements")
    total += count
    
    # ── 2. MISE A JOUR WINDOWS ── 
    # Header row 7, cols: Personnes, TYPES, Nettoyage, Fichiers temp, MAJ Windows
    ws = wb['MISE A JOUR WINDOWS']
    count = 0
    for row in ws.iter_rows(min_row=8, values_only=True):
        personne  = val(row, 0)
        types     = val(row, 1)
        nettoyage = val(row, 2)
        fichiers  = val(row, 3)
        maj_win   = val(row, 4)
        if not personne or personne.lower().startswith('pour sub'):
            continue
        nom = f"{types or 'PC'} - {personne}"
        desc = f"Nettoyage: {nettoyage or '-'} | Fichiers temp: {fichiers or '-'} | MAJ Windows: {maj_win or '-'}"
        conn.execute(text(
            "INSERT INTO equipements (site_id, sous_site, nom, famille, utilisateur_nom, description, type_equipement, is_active) "
            "VALUES (:sid, 'MISE A JOUR', :nom, :fam, :user, :desc, 'PC', 1)"
        ), {'sid': site_id, 'nom': nom, 'fam': types, 'user': personne, 'desc': desc})
        count += 1
    conn.commit()
    print(f"  MISE A JOUR WINDOWS: {count} entrées")
    total += count
    
    # ── 3. IMP ET MFP RESEAUX ── 
    # Header row 7, cols: (vide), TYPE, MARQUE, MODELE, S/N
    ws = wb['IMP ET MFP RESEAUX']
    count = 0
    for row in ws.iter_rows(min_row=8, values_only=True):
        type_eq = val(row, 1)
        marque  = val(row, 2)
        modele  = val(row, 3)
        serie   = val(row, 4)
        if not type_eq or type_eq.lower().startswith('pour sub'):
            continue
        nom = f"{type_eq} {marque or ''} {modele or ''}".strip()
        conn.execute(text(
            "INSERT INTO equipements (site_id, sous_site, nom, famille, marque, modele, numero_serie, type_equipement, is_active) "
            "VALUES (:sid, 'IMP ET MFP RESEAUX', :nom, :fam, :marq, :mod, :ser, 'IMPRIMANTE', 1)"
        ), {'sid': site_id, 'nom': nom, 'fam': type_eq, 'marq': marque, 'mod': modele, 'ser': serie})
        count += 1
    conn.commit()
    print(f"  IMP ET MFP RESEAUX: {count} équipements")
    total += count
    
    # ── 4. DATA CENTER ── 
    # Header row 7, cols: (vide), Type, Marque, Modèle, N°de série
    ws = wb['DATA CENTER']
    count = 0
    for row in ws.iter_rows(min_row=8, values_only=True):
        type_eq = val(row, 1)
        marque  = val(row, 2)
        modele  = val(row, 3)
        serie   = val(row, 4)
        if not type_eq or type_eq.lower().startswith('pour sub'):
            continue
        nom = f"{type_eq} {marque or ''}".strip()
        teq = 'RESEAU' if any(k in str(type_eq).upper() for k in ('SERVEUR', 'SWITCH', 'KVM')) else 'AUTRE'
        conn.execute(text(
            "INSERT INTO equipements (site_id, sous_site, nom, famille, marque, modele, numero_serie, type_equipement, is_active) "
            "VALUES (:sid, 'DATA CENTER', :nom, :fam, :marq, :mod, :ser, :teq, 1)"
        ), {'sid': site_id, 'nom': nom, 'fam': type_eq, 'marq': marque, 'mod': modele, 'ser': serie, 'teq': teq})
        count += 1
    conn.commit()
    print(f"  DATA CENTER: {count} équipements")
    total += count
    
    # Update feuilles
    feuilles = ['PC', 'MISE A JOUR', 'IMP ET MFP RESEAUX', 'DATA CENTER']
    conn.execute(text("UPDATE sites SET feuilles = :f WHERE id = :sid"),
                 {'f': json.dumps(feuilles), 'sid': site_id})
    conn.commit()
    print(f"  => Total Rabat: {total} entrées | Feuilles: {feuilles}")


# ── MAIN ──────────────────────────────────────────────────────────────────────
with engine.connect() as conn:
    print("=== Import AMEE Marrakech ===")
    import_marrakech(conn)
    
    print("\n=== Import AMEE Rabat ===")
    import_rabat(conn)

print("\nImport termine avec succes!")
