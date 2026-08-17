# coding: utf-8
import sys
import openpyxl
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(
    r'C:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\AMEE\MP AMEE MARRAKECH 1T-2026.xlsx',
    data_only=True
)

print("=== Feuil1 (ignorée) ===")
ws = wb['Feuil1']
print(f"Lignes: {ws.max_row} | Colonnes: {ws.max_column}")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=35, values_only=True), 1):
    vals = [str(v)[:25] if v is not None else '' for v in row[:10]]
    if any(v.strip() for v in vals):
        print(f"  L{i:2d}: {' | '.join(vals)}")

print("\n=== Feuil9 (ignorée) ===")
ws = wb['Feuil9']
print(f"Lignes: {ws.max_row}")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
    vals = [str(v)[:25] if v is not None else '' for v in row[:6]]
    if any(v.strip() for v in vals):
        print(f"  L{i:2d}: {' | '.join(vals)}")

print("\n=== UC - quelques lignes pour voir utilisateur_nom ===")
ws = wb['UC ']
for i, row in enumerate(ws.iter_rows(min_row=6, max_row=12, values_only=True), 6):
    vals = [str(v)[:20] if v is not None else '' for v in row[:9]]
    print(f"  L{i:2d}: {' | '.join(vals)}")

print("\n=== IMPRIMANTE ET MFP - lignes complètes ===")
ws = wb['IMPRIMANTE ET MFP ']
for i, row in enumerate(ws.iter_rows(min_row=7, max_row=16, values_only=True), 7):
    vals = [str(v)[:25] if v is not None else '' for v in row[:8]]
    if any(v.strip() for v in vals):
        print(f"  L{i:2d}: {' | '.join(vals)}")
