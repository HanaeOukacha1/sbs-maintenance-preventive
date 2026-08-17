import openpyxl

files = [
    r'C:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\AMEE\MP AMEE MARRAKECH 1T-2026.xlsx',
    r'C:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\AMEE\NV MP AMEE RABAT 1T-2026.xlsx',
]

for f in files:
    print(f'\n\n========== {f.split(chr(92))[-1]} ==========')
    wb = openpyxl.load_workbook(f, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'\n  --- Feuille: {sheet_name} ---')
        # Show first 5 rows max, 15 cols max
        for row in ws.iter_rows(min_row=1, max_row=8, max_col=20, values_only=True):
            row_vals = [str(v)[:25] if v is not None else '' for v in row]
            if any(v for v in row_vals):
                print('  ', ' | '.join(row_vals))
