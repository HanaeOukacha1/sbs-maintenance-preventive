import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\ONP\MP ONP.xlsx', data_only=True)
print("Sheets:", wb.sheetnames)

for sheet_name in wb.sheetnames[:3]:
    ws = wb[sheet_name]
    print(f"\n=== {sheet_name} ===")
    for row in ws.iter_rows(min_row=1, max_row=10, max_col=15, values_only=True):
        row_vals = [str(v)[:20] if v is not None else '' for v in row]
        if any(v for v in row_vals):
            print(' | '.join(row_vals))
