# coding: utf-8
from sqlalchemy import create_engine, text
import openpyxl

engine = create_engine('mysql+pymysql://root:@localhost/sbs_db', pool_pre_ping=True)
with engine.connect() as conn:
    sites = conn.execute(text("SELECT id, nom, checklist_type, feuilles FROM sites WHERE checklist_type = 'MHAI'")).fetchall()
    print('MHAI sites:')
    for s in sites:
        count = conn.execute(text(f"SELECT COUNT(*) FROM equipements WHERE site_id = {s[0]}")).fetchone()
        missions = conn.execute(text(f"SELECT id, titre, technicien_id FROM missions WHERE site_id = {s[0]}")).fetchall()
        print(f"  Site {s[0]}: {s[1]} | EQ: {count[0]} | Missions: {missions}")

# Check Excel
wb = openpyxl.load_workbook(
    r'C:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\MHAI\MP HABOUS S2-24.xlsx',
    data_only=True
)
print('\nExcel sheets:', wb.sheetnames)
for sheet_name in wb.sheetnames[:3]:
    ws = wb[sheet_name]
    print(f'\n=== {sheet_name} ===')
    for row in ws.iter_rows(min_row=1, max_row=8, max_col=12, values_only=True):
        row_vals = [str(v)[:20] if v is not None else '' for v in row]
        if any(v for v in row_vals):
            print('  ', ' | '.join(row_vals))
