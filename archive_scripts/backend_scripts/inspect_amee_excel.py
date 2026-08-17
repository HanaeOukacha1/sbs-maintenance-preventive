# coding: utf-8
import openpyxl

for fname in ['MP AMEE MARRAKECH 1T-2026.xlsx', 'NV MP AMEE RABAT 1T-2026.xlsx']:
    wb = openpyxl.load_workbook(
        rf'C:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\AMEE\{fname}',
        data_only=True
    )
    print(f"\n{'='*50}")
    print(f"File: {fname}")
    print(f"Sheets: {wb.sheetnames}")
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        print(f"\n  -- Sheet: {sheet} --")
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            vals = [str(v)[:20] if v else '' for v in row[:10]]
            if any(vals):
                print("   ", " | ".join(vals))
        # Count rows with data
        count = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if any(v for v in r if v))
        print(f"   => ~{count} lignes de données")
