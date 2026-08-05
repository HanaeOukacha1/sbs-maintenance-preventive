# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.drawing.image import Image
from io import BytesIO
import os
import base64
from copy import copy

def export_msante_capm(mission, interventions, equipements):
    template_path = os.path.join(os.path.dirname(__file__), '..', 'templates', 'template_msante_capm.xlsx')
    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active
    
    start_row = 10
    ref_row = 10
    
    for idx, eq in enumerate(equipements):
        row_idx = start_row + idx
        intervention = next((i for i in interventions if i.equipement_id == eq.id), None)
        
        designation = getattr(eq, 'designation', None) or getattr(eq, 'famille', None) or getattr(eq, 'type_equipement', None) or ""
        marque = getattr(eq, 'marque', None) or ""
        modele = getattr(eq, 'modele', None) or ""
        n_serie = getattr(eq, 'numero_serie', None) or ""
        utilisateur = getattr(eq, 'utilisateur_nom', None) or ""
        signature_base64 = None
        
        if intervention:
            reponses = intervention.reponses or {}
            eq_mod = reponses.get("equipement_modifie", {})
            designation = eq_mod.get("designation") or designation
            marque = eq_mod.get("marque") or marque
            modele = eq_mod.get("modele") or modele
            n_serie = eq_mod.get("numero_serie") or n_serie
            utilisateur = eq_mod.get("utilisateur") or utilisateur
            signature_base64 = reponses.get("signature")
            
            if intervention.est_hors_inventaire and intervention.equipement_hors_inventaire:
                eq_hi = intervention.equipement_hors_inventaire
                designation = eq_hi.get('designation') or designation
                marque = eq_hi.get('marque') or marque
                modele = eq_hi.get('modele') or modele
                n_serie = eq_hi.get('numero_serie') or n_serie
        
        sheet.cell(row=row_idx, column=1, value=idx+1)
        sheet.cell(row=row_idx, column=2, value=designation)
        sheet.cell(row=row_idx, column=3, value=marque)
        sheet.cell(row=row_idx, column=4, value=utilisateur)
        # column 5 is Signature (image)
        sheet.cell(row=row_idx, column=6, value=modele)
        sheet.cell(row=row_idx, column=7, value=n_serie)
        
        # Add signature image
        if signature_base64 and ',' in signature_base64:
            try:
                img_data = base64.b64decode(signature_base64.split(',')[1])
                img_io = BytesIO(img_data)
                img = Image(img_io)
                img.height = 40
                img.width = 120
                col_letter = openpyxl.utils.get_column_letter(5)
                sheet.add_image(img, f'{col_letter}{row_idx}')
                sheet.row_dimensions[row_idx].height = 45
            except Exception as e:
                print(f"Error adding signature: {e}")
                pass
        
        for c in range(1, 8):
            cell = sheet.cell(row=ref_row, column=c)
            new_cell = sheet.cell(row=row_idx, column=c)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.alignment = copy(cell.alignment)
            
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"Rapport_MSANTE_CAPM_{mission.site.nom if mission.site else ''}.xlsx"

def export_msante_dprf(mission, interventions, equipements):
    template_path = os.path.join(os.path.dirname(__file__), '..', 'templates', 'template_msante_dprf.xlsx')
    wb = openpyxl.load_workbook(template_path)
    
    # We group equipements by sous_site (feuille)
    from collections import defaultdict
    sheets_data = defaultdict(list)
    for eq in equipements:
        feuille = getattr(eq, 'sous_site', None) or 'Nouveau'
        sheets_data[feuille].append(eq)
        
    for sheet_name, eqs in sheets_data.items():
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            # If sheet doesn't exist, duplicate the first one
            sheet = wb.copy_worksheet(wb.worksheets[0])
            sheet.title = str(sheet_name)[:31]
            
        start_row = 10
        ref_row = 10
        
        for idx, eq in enumerate(eqs):
            row_idx = start_row + idx
            intervention = next((i for i in interventions if i.equipement_id == eq.id), None)
            
            designation = getattr(eq, 'designation', None) or getattr(eq, 'famille', None) or getattr(eq, 'type_equipement', None) or ""
            marque = getattr(eq, 'marque', None) or ""
            modele = getattr(eq, 'modele', None) or ""
            n_serie = getattr(eq, 'numero_serie', None) or ""
            utilisateur = getattr(eq, 'utilisateur_nom', None) or ""
            signature_base64 = None
            
            if intervention:
                reponses = intervention.reponses or {}
                eq_mod = reponses.get("equipement_modifie", {})
                designation = eq_mod.get("designation") or designation
                marque = eq_mod.get("marque") or marque
                modele = eq_mod.get("modele") or modele
                n_serie = eq_mod.get("numero_serie") or n_serie
                utilisateur = eq_mod.get("utilisateur") or utilisateur
                signature_base64 = reponses.get("signature")
                
                if intervention.est_hors_inventaire and intervention.equipement_hors_inventaire:
                    eq_hi = intervention.equipement_hors_inventaire
                    designation = eq_hi.get('designation') or designation
                    marque = eq_hi.get('marque') or marque
                    modele = eq_hi.get('modele') or modele
                    n_serie = eq_hi.get('numero_serie') or n_serie
            
            # DPRF Order: N°, Désignation, Utilisateurs, signature, Marque, Article, N° série
            sheet.cell(row=row_idx, column=1, value=idx+1)
            sheet.cell(row=row_idx, column=2, value=designation)
            sheet.cell(row=row_idx, column=3, value=utilisateur)
            # col 4: signature
            sheet.cell(row=row_idx, column=5, value=marque)
            sheet.cell(row=row_idx, column=6, value=modele)
            sheet.cell(row=row_idx, column=7, value=n_serie)
            
            if signature_base64 and ',' in signature_base64:
                try:
                    img_data = base64.b64decode(signature_base64.split(',')[1])
                    img_io = BytesIO(img_data)
                    img = Image(img_io)
                    img.height = 40
                    img.width = 120
                    col_letter = openpyxl.utils.get_column_letter(4)
                    sheet.add_image(img, f'{col_letter}{row_idx}')
                    sheet.row_dimensions[row_idx].height = 45
                except Exception:
                    pass
            
            for c in range(1, 8):
                cell = sheet.cell(row=ref_row, column=c)
                new_cell = sheet.cell(row=row_idx, column=c)
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.alignment = copy(cell.alignment)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"Rapport_MSANTE_DPRF_{mission.site.nom if mission.site else ''}.xlsx"
