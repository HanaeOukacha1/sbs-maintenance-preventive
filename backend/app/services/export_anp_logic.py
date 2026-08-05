import openpyxl
from io import BytesIO
import os
from copy import copy

def copy_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.border = copy(src_cell.border)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.number_format = copy(src_cell.number_format)
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.alignment = copy(src_cell.alignment)

def export_anp(mission, interventions, equipements):
    template_path = os.path.join(os.path.dirname(__file__), '..', 'templates', 'template_anp.xlsx')
    wb = openpyxl.load_workbook(template_path)
    sheet = wb.active
    
    if mission.site:
        sheet['A5'] = f"Site: {mission.site.nom}"
        
    start_row = 7
    ref_row = 6
    
    for idx, eq in enumerate(equipements):
        row_idx = start_row + idx
        intervention = next((i for i in interventions if i.equipement_id == eq.id), None)
        
        designation = getattr(eq, 'designation', None) or getattr(eq, 'famille', None) or getattr(eq, 'type_equipement', None) or ""
        marque = getattr(eq, 'marque', None) or ""
        modele = getattr(eq, 'modele', None) or ""
        n_serie = getattr(eq, 'numero_serie', None) or ""
        etat = ""
        
        if intervention:
            reponses = intervention.reponses or {}
            eq_mod = reponses.get("equipement_modifie", {})
            designation = eq_mod.get("designation") or designation
            marque = eq_mod.get("marque") or marque
            modele = eq_mod.get("modele") or modele
            n_serie = eq_mod.get("numero_serie") or n_serie

            if intervention.est_hors_inventaire and intervention.equipement_hors_inventaire:
                eq_hi = intervention.equipement_hors_inventaire
                designation = eq_hi.get('designation') or designation
                marque = eq_hi.get('marque') or marque
                modele = eq_hi.get('modele') or modele
                n_serie = eq_hi.get('numero_serie') or n_serie
            
            etat = reponses.get("etat") or ""
        
        sheet.cell(row=row_idx, column=1, value=idx+1)
        sheet.cell(row=row_idx, column=2, value=designation)
        sheet.cell(row=row_idx, column=3, value=marque)
        sheet.cell(row=row_idx, column=4, value=modele)
        sheet.cell(row=row_idx, column=5, value=n_serie)
        sheet.cell(row=row_idx, column=6, value=etat)
        
        for c in range(1, 7):
            copy_style(sheet.cell(row=ref_row, column=c), sheet.cell(row=row_idx, column=c))
            
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"Rapport_ANP_{mission.site.nom if mission.site else ''}.xlsx"
