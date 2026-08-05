# coding: utf-8
import sys
import openpyxl
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(
    r'C:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\AMEE\MP AMEE MARRAKECH 1T-2026.xlsx',
    data_only=True
)

print(f"Feuilles disponibles: {wb.sheetnames}")
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== {sheet_name} ===")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        vals = [str(v)[:25] if v is not None else '' for v in row[:10]]
        if any(v.strip() for v in vals):
            print(f"  L{i:2d}: {' | '.join(vals)}")
