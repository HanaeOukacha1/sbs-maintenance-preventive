# coding: utf-8
from sqlalchemy import create_engine, text
import openpyxl
import json

engine = create_engine('mysql+pymysql://root:@localhost/sbs_db', pool_pre_ping=True)

# ── Config par site AMEE ──────────────────────────────────────────────────────
AMEE_CONFIG = [
    {
        'site_nom': 'AMEE Marrakech',
        'checklist_type': 'AMEE_MARRAKECH',
        'excel_file': r'C:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\AMEE\MP AMEE MARRAKECH 1T-2026.xlsx',
        # nom_feuille_excel -> nom affiché dans l'app (sous_site)
        'sheets_map': {
            'DATA CENTER': 'DATA CENTER',
            'UC ': 'UC',
            'MISE A JOUR': 'MISE A JOUR',
            'IMPRIMANTE ET MFP ': 'IMPRIMANTE ET MFP',
        },
        # header col positions pour chaque feuille (cherché dynamiquement)
        # Les colonnes clés attendues par feuille
        'fields_by_sheet': {
            'DATA CENTER': {'type': 4, 'marque': 5, 'modele': 6, 'serie': 7},
            'UC': {'utilisateur': None, 'type': None, 'marque': None, 'modele': None, 'serie': None, 'cpu': None, 'ram': None},
            'MISE A JOUR': {'type': None, 'marque': None, 'modele': None, 'serie': None},
            'IMPRIMANTE ET MFP': {'type': None, 'marque': None, 'modele': None, 'serie': None},
        }
    },
    {
        'site_nom': 'AMEE Rabat',
        'checklist_type': 'AMEE_RABAT',
        'excel_file': r'C:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\AMEE\NV MP AMEE RABAT 1T-2026.xlsx',
        'sheets_map': {
            'PC.': 'PC',
            'MISE A JOUR WINDOWS': 'MISE A JOUR',
            'IMP ET MFP RESEAUX': 'IMP ET MFP RESEAUX',
            'DATA CENTER': 'DATA CENTER',
        },
    },
]

def find_header_row(ws, col_keywords=['marque', 'modele', 'serie', 'utilisateur', 'type', 'designation', 'n° de']):
    """Find the row index where column headers are located"""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True)):
        vals = [str(x).lower().strip() if x else '' for x in row]
        hits = sum(1 for v in vals if any(k in v for k in col_keywords))
        if hits >= 2:
            return i + 1  # 1-indexed
    return None

def map_columns(ws, header_row):
    """Map column names to indices"""
    row = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    col_map = {}
    for idx, val in enumerate(row):
        if val is None:
            continue
        v = str(val).lower().strip()
        if 'utilisateur' in v: col_map['utilisateur'] = idx
        elif 'type' in v or 'désignation' in v or 'designation' in v: 
            if 'type' not in col_map: col_map['type'] = idx
        elif 'marque' in v: col_map['marque'] = idx
        elif 'modele' in v or 'modèle' in v: col_map['modele'] = idx
        elif 'n° de serie' in v or 'n° série' in v or 'serie' in v or 'série' in v: 
            if 'serie' not in col_map: col_map['serie'] = idx
        elif 'cpu' in v or 'processeur' in v: col_map['cpu'] = idx
        elif 'ram' in v or 'mémoire' in v: col_map['ram'] = idx
        elif 'n° inv' in v or 'inventaire' in v: col_map['inventaire'] = idx
        elif 'sys' in v or 'os' in v: col_map['os'] = idx
    return col_map

def get_val(row_vals, col_map, key):
    if key in col_map:
        idx = col_map[key]
        if idx < len(row_vals):
            v = row_vals[idx]
            if v is not None and str(v).strip().lower() not in ('', 'none', 'nan'):
                return str(v).strip()[:150]
    return None

with engine.connect() as conn:
    for config in AMEE_CONFIG:
        site_nom = config['site_nom']
        site = conn.execute(text("SELECT id FROM sites WHERE nom = :nom"), {'nom': site_nom}).fetchone()
        if not site:
            print(f"Site {site_nom} not found in DB")
            continue
        site_id = site[0]
        
        # Clear existing equipements and reinsert
        conn.execute(text(f"DELETE FROM equipements WHERE site_id = {site_id}"))
        conn.commit()
        print(f"\n=== {site_nom} (id={site_id}) ===")
        
        # Determine feuilles list for the site
        feuilles_list = list(config['sheets_map'].values())
        conn.execute(text("UPDATE sites SET feuilles = :f WHERE id = :sid"), 
                     {'f': json.dumps(feuilles_list), 'sid': site_id})
        conn.commit()
        print(f"  Feuilles: {feuilles_list}")
        
        wb = openpyxl.load_workbook(config['excel_file'], data_only=True)
        
        total_eq = 0
        for excel_sheet, sous_site_name in config['sheets_map'].items():
            if excel_sheet not in wb.sheetnames:
                print(f"  Sheet '{excel_sheet}' not in Excel, skipping")
                continue
            ws = wb[excel_sheet]
            
            header_row = find_header_row(ws)
            if header_row is None:
                print(f"  Sheet '{excel_sheet}': no header found, skipping")
                continue
            
            col_map = map_columns(ws, header_row)
            print(f"  Sheet '{excel_sheet}' -> sous_site='{sous_site_name}': header@row{header_row}, cols={list(col_map.keys())}")
            
            inserted = 0
            for row in ws.iter_rows(min_row=header_row+1, values_only=True):
                if not any(v for v in row if v is not None):
                    continue
                
                type_val = get_val(row, col_map, 'type')
                marque = get_val(row, col_map, 'marque')
                modele = get_val(row, col_map, 'modele')
                serie = get_val(row, col_map, 'serie')
                
                if not type_val and not marque and not serie:
                    continue
                
                utilisateur = get_val(row, col_map, 'utilisateur')
                cpu = get_val(row, col_map, 'cpu')
                ram = get_val(row, col_map, 'ram')
                
                type_eq = 'PC'
                if type_val:
                    tl = type_val.lower()
                    if 'imprimante' in tl or 'mfp' in tl or 'scanner' in tl:
                        type_eq = 'IMPRIMANTE'
                    elif 'serveur' in tl or 'switch' in tl or 'baie' in tl:
                        type_eq = 'RESEAU'
                
                conn.execute(text(
                    "INSERT INTO equipements (site_id, sous_site, utilisateur_nom, famille, marque, modele, numero_serie, cpu, ram, type_equipement, is_active) "
                    "VALUES (:sid, :ss, :user, :fam, :marq, :mod, :ser, :cpu, :ram, :teq, 1)"
                ), {
                    'sid': site_id, 'ss': sous_site_name, 'user': utilisateur,
                    'fam': type_val, 'marq': marque, 'mod': modele, 'ser': serie,
                    'cpu': cpu, 'ram': ram, 'teq': type_eq
                })
                inserted += 1
            
            conn.commit()
            total_eq += inserted
            print(f"    => Inserted {inserted} equipements")
        
        print(f"  Total: {total_eq} equipements for {site_nom}")

print("\nDone!")
