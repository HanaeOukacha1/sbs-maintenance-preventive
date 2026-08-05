# coding: utf-8
import openpyxl

wb = openpyxl.load_workbook(
    r'C:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\MARSA MAROC\MP MM .xlsx',
    data_only=True
)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n=== Feuille: {sheet_name} ===')
    for row in ws.iter_rows(min_row=1, max_row=12, max_col=20, values_only=True):
        row_vals = [str(v)[:22] if v is not None else '' for v in row]
        if any(v for v in row_vals):
            print('  ', ' | '.join(row_vals))
