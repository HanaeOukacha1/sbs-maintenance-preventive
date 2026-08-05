import openpyxl
import sys
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\ANP\MP ANP.xlsx')
for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f"Sheet: {sheet}")
    for row in ws.iter_rows(values_only=True, min_row=1, max_row=15):
        non_empty = [str(x) for x in row if x is not None]
        if non_empty:
            print(non_empty)
