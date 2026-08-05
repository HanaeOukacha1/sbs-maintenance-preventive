# coding: utf-8
from sqlalchemy import create_engine, text
import openpyxl
import json

engine = create_engine('mysql+pymysql://root:@localhost/sbs_db', pool_pre_ping=True)
wb = openpyxl.load_workbook(r'C:\Users\hanae\Desktop\Stage PFA 2026\MASTER DATA\MHAI\MP HABOUS S2-24.xlsx', data_only=True)

with engine.connect() as conn:
    # Get the MHAI Tanger site
    site = conn.execute(text("SELECT id, nom FROM sites WHERE nom = 'MHAI Tanger'")).fetchone()
    if not site:
        print("Site MHAI Tanger not found")
        exit()
    site_id = site[0]
    print(f"Site: {site}")
    
    # Find what sheet corresponds to Tanger in the Excel
    print("\nAll sheets in Excel:", wb.sheetnames)
    tanger_sheets = [s for s in wb.sheetnames if 'tang' in s.lower()]
    print("Tanger sheets:", tanger_sheets)
    
    # Look at the sheets with data 
    for s in wb.sheetnames[:5]:
        ws = wb[s]
        max_row = ws.max_row
        print(f"\n{s}: {max_row} rows")
        # Look at first row
        for row in ws.iter_rows(min_row=1, max_row=2, values_only=True):
            print("  ", [str(v)[:15] if v else '' for v in row[:8]])
